###用于训练模型的初步选择
import openpyxl
import pandas as pd
import numpy as np
from sklearn.ensemble import (RandomForestRegressor, GradientBoostingRegressor,BaggingRegressor,AdaBoostRegressor,ExtraTreesRegressor)
from sklearn.svm import SVR
from sklearn.tree import DecisionTreeRegressor
from sklearn.model_selection import train_test_split
from sklearn.metrics import explained_variance_score, mean_absolute_error, r2_score,mean_squared_error
from sklearn.neural_network import MLPRegressor
from sklearn.preprocessing import StandardScaler
from statsmodels.tsa.ar_model import AR
from sklearn.linear_model import (Ridge,Lasso,ARDRegression,LinearRegression,LogisticRegression,PassiveAggressiveRegressor,PoissonRegressor,RANSACRegressor,SGDRegressor,TheilSenRegressor,TweedieRegressor)
from xgboost import XGBRegressor


#1.机器学习算法罗列
rf_model=RandomForestRegressor()  #1
GBDT_model=GradientBoostingRegressor()#1
bag_model=BaggingRegressor()
ada_model=AdaBoostRegressor()
extra_model=ExtraTreesRegressor()
svr_model=SVR()
MLP_model=MLPRegressor()#1
ridge_model=Ridge()#1
lasso_model=Lasso()
ARD_model=ARDRegression()#1
LR_model=LinearRegression()#1
PAR_model=PassiveAggressiveRegressor()#1
RSR_model=RANSACRegressor()#1
Thei_model=TheilSenRegressor()#1
Twee_model=TweedieRegressor()
xgb_model=XGBRegressor()


#2.读取数据
p=18#回归模型阶数
sample = np.array(pd.read_excel('filename.xlsx'))#70096
index = 17502 # 52553验证从哪里开始
train_x = sample[0:index, 0:p]
test_x = sample[index:, 0:p]
train_y = sample[0:index, -1]
test_y = sample[index:, -1]

#3.验证模型测试效果
allpredict_y=[]
for clf in [rf_model, GBDT_model, bag_model,ada_model,extra_model,svr_model,MLP_model,ridge_model,lasso_model,ARD_model,LR_model,PAR_model,RSR_model,Thei_model,Twee_model,xgb_model]:
    print(clf)
    clf.fit(train_x, train_y)
    predict_y = pd.Series(clf.predict(test_x))
    allpredict_y.append(predict_y)
allpredict_y = pd.DataFrame(allpredict_y).T
columnsname=['rf_model', 'GBDT_model', 'bag_model','ada_model','extra_model','svr_model','MLP_model','ridge_model','lasso_model','ARD_model','LR_model','PAR_model','RSR_model','Thei_model','Twee_model','xgb_model']
allpredict_y.columns=columnsname
res = pd.DataFrame({'real_label': pd.Series(test_y).map(lambda x: float(x)).tolist()})
evs=[]
mae=[]
mse=[]
r2score=[]
for i in range(len(columnsname)):
    predict=allpredict_y[columnsname[i]].tolist()
    evs.append(explained_variance_score(res['real_label'], predict))
    r2score.append(r2_score(res['real_label'], predict))
    mae.append(mean_absolute_error(res['real_label'],  predict))
    mse.append(mean_squared_error(res['real_label'],  predict))
indicatorvalue=[]
indicatorvalue=[evs,mae,r2score,mse]

#4.结果导出
wb = openpyxl.load_workbook('filename.xlsx')#计算结果导出excel
sheetname = 'MLselected'
ws = wb.create_sheet(sheetname)
for i in range(len(evs)):
    for j in range(4):#4个指标
        if j == 0:
            ws.cell(row=i + 1, column=1).value = columnsname[i]
        ws.cell(row=i + 1, column=2+j).value = indicatorvalue[j][i]
wb.save('filename.xlsx')
