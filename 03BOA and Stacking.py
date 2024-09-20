#####超参数优化
# 机器学习集成进行实时校正

import pandas as pd
import numpy as np
from sklearn.model_selection import KFold
from sklearn.ensemble import (RandomForestRegressor, GradientBoostingRegressor,ExtraTreesRegressor)
from sklearn.svm import SVR
from sklearn.tree import DecisionTreeRegressor
from sklearn.model_selection import train_test_split
from sklearn.metrics import explained_variance_score, mean_absolute_error, r2_score,mean_squared_error
from sklearn.neural_network import MLPRegressor
from sklearn.preprocessing import StandardScaler
from statsmodels.tsa.ar_model import AR
from sklearn.linear_model import (Ridge,Lasso,ARDRegression,GammaRegressor,LinearRegression,LogisticRegression,PassiveAggressiveRegressor,PoissonRegressor,RANSACRegressor,SGDRegressor,TheilSenRegressor,TweedieRegressor)
import openpyxl
from xgboost import XGBRegressor
#from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score#分类任务的评价标准
import matplotlib.pyplot as plt
from bayes_opt import BayesianOptimization
import statistics



###############贝叶斯优化####################################
#函数输入为机器学习算法的所有超参数。因为bayes_opt库只支持最大值，所以最后的输出如果是越小越好，那么需要在前面加上负号，以转为最大值！！！！
#RF
def rf_cv(n_estimators, min_samples_split, max_depth):
    model = RandomForestRegressor(n_estimators=int(n_estimators),
                                  min_samples_split=int(min_samples_split),
                                  max_depth=int(max_depth))
    r2score=[]
    for i, (train_index, test_index) in enumerate(kf.split(train_x)):
        ## 用来建模的部分
        x_tra, y_tra = pd.DataFrame(train_x).iloc[train_index], pd.DataFrame(train_y).iloc[train_index]
        ## 用来做验证的部分
        x_tst, y_tst = pd.DataFrame(train_x).iloc[test_index], pd.DataFrame(train_y).iloc[test_index]
        y_tra=np.array(y_tra).ravel()#将二维数组转化成一维数组
        model.fit(x_tra, y_tra)
        predict=model.predict(x_tst)
        r2score.append(r2_score(y_tst, predict))
    return statistics.mean(r2score)
#gbdt
def gbdt_cv(n_estimators, min_samples_split, max_depth):
    model = GradientBoostingRegressor(n_estimators=int(n_estimators),
                                  min_samples_split=int(min_samples_split),
                                  max_depth=int(max_depth))
    r2score=[]
    for i, (train_index, test_index) in enumerate(kf.split(train_x)):
        ## 用来建模的部分
        x_tra, y_tra = pd.DataFrame(train_x).iloc[train_index], pd.DataFrame(train_y).iloc[train_index]
        ## 用来做验证的部分
        x_tst, y_tst = pd.DataFrame(train_x).iloc[test_index], pd.DataFrame(train_y).iloc[test_index]
        y_tra=np.array(y_tra).ravel()
        model.fit(x_tra, y_tra)
        predict=model.predict(x_tst)
        r2score.append(r2_score(y_tst, predict))
    return statistics.mean(r2score)
#MLPRegressor
def MLP_cv(hidden_layer_sizes,max_iter):
    model =MLPRegressor(hidden_layer_sizes=int(hidden_layer_sizes),max_iter=int(max_iter))
    r2score = []
    for i, (train_index, test_index) in enumerate(kf.split(train_x)):
        ## 用来建模的部分
        x_tra, y_tra = pd.DataFrame(train_x).iloc[train_index], pd.DataFrame(train_y).iloc[train_index]
        ## 用来做验证的部分
        x_tst, y_tst = pd.DataFrame(train_x).iloc[test_index], pd.DataFrame(train_y).iloc[test_index]
        y_tra = np.array(y_tra).ravel()
        model.fit(x_tra, y_tra)
        predict = model.predict(x_tst)
        r2score.append(r2_score(y_tst, predict))
    return statistics.mean(r2score)
#ridge
def ridge_cv(alpha):
    model =Ridge(alpha=alpha)
    r2score = []
    for i, (train_index, test_index) in enumerate(kf.split(train_x)):
        ## 用来建模的部分
        x_tra, y_tra = pd.DataFrame(train_x).iloc[train_index], pd.DataFrame(train_y).iloc[train_index]
        ## 用来做验证的部分
        x_tst, y_tst = pd.DataFrame(train_x).iloc[test_index], pd.DataFrame(train_y).iloc[test_index]
        y_tra = np.array(y_tra).ravel()
        model.fit(x_tra, y_tra)
        predict = model.predict(x_tst)
        r2score.append(r2_score(y_tst, predict))
    return statistics.mean(r2score)
#ARDRegressor
def ard_cv(alpha_1, alpha_2, lambda_1,lambda_2):
    model=ARDRegression(alpha_1=alpha_1,alpha_2=alpha_2,
                        lambda_1=lambda_1,lambda_2=lambda_2)
    r2score = []
    for i, (train_index, test_index) in enumerate(kf.split(train_x)):
        ## 用来建模的部分
        x_tra, y_tra = pd.DataFrame(train_x).iloc[train_index], pd.DataFrame(train_y).iloc[train_index]
        ## 用来做验证的部分
        x_tst, y_tst = pd.DataFrame(train_x).iloc[test_index], pd.DataFrame(train_y).iloc[test_index]
        y_tra = np.array(y_tra).ravel()
        model.fit(x_tra, y_tra)
        predict = model.predict(x_tst)
        r2score.append(r2_score(y_tst, predict))
    return statistics.mean(r2score)

#SVR
def svr_cv(C, epsilon):
    model=SVR(kernel='rbf', C=float(C),epsilon=epsilon)
    model.fit(train_x, train_y)
    predict = model.predict(test_x)
    val = r2_score(test_y, predict)
    return val
#ExtraTreesRegressor
def et_cv(n_estimators, min_samples_split, max_depth):
    model=ExtraTreesRegressor(n_estimators=int(n_estimators),
                        min_samples_split=int(min_samples_split),
                        max_depth=int(max_depth))
    model.fit(train_x, train_y)
    predict = model.predict(test_x)
    val = r2_score(test_y, predict)
    return val

###################get_stacking函数#####################
#param:clf：基分类器；x_train：训练集特征；y_train：训练集标签；x_test：测试集特征；n_folds：折叠交叉次数
#result:second_level_train_set:用clf预测的训练集，为第二层模型构建提供的数据；second_level_test_set：用clf预测的测试集，为第二层模型的测试集提供的数据
def get_stacking(clf, x_train, y_train, x_test, n_folds):
    """
    这个函数是stacking的核心，使用交叉验证的方法得到次级训练集
    x_train, y_train, x_test 的值应该为numpy里面的数组类型 numpy.ndarray .
    如果输入为pandas的DataFrame类型则会把报错
    """
    train_num, test_num = x_train.shape[0], x_test.shape[0]
    second_level_train_set = np.zeros((train_num,))
    second_level_test_set = np.zeros((test_num,))
    test_nfolds_sets = np.zeros((test_num, n_folds))
    kf = KFold(n_splits=n_folds)

    for i, (train_index, test_index) in enumerate(kf.split(x_train)):
        ## 用来建模的部分
        x_tra, y_tra = pd.DataFrame(x_train).iloc[train_index], pd.DataFrame(y_train).iloc[train_index]
        ## 用来做验证的部分
        x_tst, y_tst = pd.DataFrame(x_train).iloc[test_index], pd.DataFrame(y_train).iloc[test_index]
        y_tra=np.array(y_tra).ravel()#将（135,1）的二维数组转化成一维数组
        clf.fit(x_tra, y_tra)
        second_level_train_set[test_index] = clf.predict(x_tst)
        test_nfolds_sets[:, i] = clf.predict(x_test)

    second_level_test_set[:] = test_nfolds_sets.mean(axis=1)
    return second_level_train_set, second_level_test_set

n_folds=4#设置交叉验证折数
kf = KFold(n_splits=n_folds)

#输入数据
p=18#回归模型阶数
sample = np.array(pd.read_excel('filename.xlsx'))#70096
index = 17502 # 验证从哪里开始
train_x = sample[0:index, 3:3+p]
test_x = sample[index:, 3:3+p]
train_y = sample[0:index, -1]
test_y = sample[index:, -1]
Q_forecasttest = sample[index:, 0]
Q_observedtest = sample[index:, 1]
metricsave=[]

#超参数优化
##################1.RF
rf = RandomForestRegressor()
val=0
for i, (train_index, test_index) in enumerate(kf.split(train_x)):
    ## 用来建模的部分
    x_tra, y_tra = pd.DataFrame(train_x).iloc[train_index], pd.DataFrame(train_y).iloc[train_index]
    ## 用来做验证的部分
    x_tst, y_tst = pd.DataFrame(train_x).iloc[test_index], pd.DataFrame(train_y).iloc[test_index]
    y_tra = np.array(y_tra).ravel()  # 将（135,1）的二维数组转化成一维数组
    rf.fit(x_tra, y_tra)
    predict = rf.predict(x_tst)
    tt=r2_score(y_tst, predict)
    print(tt)
    val=val+r2_score(y_tst, predict)
print(val/n_folds)
# a.建立贝叶斯优化对象（第一个参数是优化目标函数，第二个参数是所需优化的超参数名称及其对应范围）
rf_bo = BayesianOptimization(
    rf_cv,
    {'n_estimators': (10, 1000),
     'min_samples_split': (2, 25),
     'max_depth': (1, 150)})
# b.开始优化
rf_bo.maximize(init_points=5, n_iter=20)
# c.寻找最优的几组参数做贝叶斯进阶优化
paramlist = ['n_estimators', 'min_samples_split',  'max_depth']
result = rf_bo.res
result.sort(key=lambda x: x['target'])  # 按照target的值从小到大排列
result.reverse()
metrics = result[0]['params']  # 选取最好的一组参数
metricsave.append(metrics)
print(metrics)
rf_model = RandomForestRegressor(n_estimators=int(metrics[paramlist[0]]),
                                 min_samples_split=int(metrics[paramlist[1]]),
                                max_depth=int(metrics[paramlist[2]]),oob_score=True)

##################2.ET
extra = ExtraTreesRegressor()
val=0
for i, (train_index, test_index) in enumerate(kf.split(train_x)):
    ## 用来建模的部分
    x_tra, y_tra = pd.DataFrame(train_x).iloc[train_index], pd.DataFrame(train_y).iloc[train_index]
    ## 用来做验证的部分
    x_tst, y_tst = pd.DataFrame(train_x).iloc[test_index], pd.DataFrame(train_y).iloc[test_index]
    y_tra = np.array(y_tra).ravel()  # 二维数组转化成一维数组
    extra.fit(x_tra, y_tra)
    predict = extra.predict(x_tst)
    tt=r2_score(y_tst, predict)
    print(tt)
    val=val+r2_score(y_tst, predict)
print(val/n_folds)
# a.建立贝叶斯优化对象（第一个参数是优化目标函数，第二个参数是所需优化的超参数名称及其对应范围）
extra_bo = BayesianOptimization(
    et_cv,
    {'n_estimators': (10, 1000),
     'min_samples_split': (2, 25),
     'max_depth': (1, 150)})
# b.开始优化
extra_bo.maximize(init_points=5, n_iter=20)
# c.寻找最优的几组参数做贝叶斯进阶优化
paramlist = ['n_estimators', 'min_samples_split',  'max_depth']
result = extra_bo.res
result.sort(key=lambda x: x['target'])  # 按照target的值从小到大排列
result.reverse()
metrics = result[0]['params']  # 选取最好的一组参数
metricsave.append(metrics)
print(metrics)
extra_model = ExtraTreesRegressor(n_estimators=int(metrics[paramlist[0]]),
                                 min_samples_split=int(metrics[paramlist[1]]),
                                max_depth=int(metrics[paramlist[2]]),oob_score=True, bootstrap=True)

##################3.gbdt
gbdt = GradientBoostingRegressor()
val=0
for i, (train_index, test_index) in enumerate(kf.split(train_x)):
    ## 用来建模的部分
    x_tra, y_tra = pd.DataFrame(train_x).iloc[train_index], pd.DataFrame(train_y).iloc[train_index]
    ## 用来做验证的部分
    x_tst, y_tst = pd.DataFrame(train_x).iloc[test_index], pd.DataFrame(train_y).iloc[test_index]
    y_tra = np.array(y_tra).ravel()  # 将（135,1）的二维数组转化成一维数组
    gbdt.fit(x_tra, y_tra)
    predict = gbdt.predict(x_tst)
    tt=r2_score(y_tst, predict)
    print(tt)
    val=val+r2_score(y_tst, predict)
print(val/n_folds)
# a.建立贝叶斯优化对象（第一个参数是优化目标函数，第二个参数是所需优化的超参数名称及其对应范围）
gbdt_bo = BayesianOptimization(
    gbdt_cv,
    {'n_estimators': (50, 500),
     'min_samples_split': (2, 20),
     'max_depth': (3, 10)})
# b.开始优化
#svr_bo.maximize(n_iter=3)
gbdt_bo.maximize(init_points=5, n_iter=20)
# c.寻找最优的几组参数做贝叶斯进阶优化
paramlist = ['n_estimators', 'min_samples_split',  'max_depth']
result = gbdt_bo.res
result.sort(key=lambda x: x['target'])  # 按照target的值从小到大排列
result.reverse()
metrics = result[0]['params']
metricsave.append(metrics)
print(metrics)
gbdt_model = GradientBoostingRegressor(n_estimators=int(metrics[paramlist[0]]),
                                 min_samples_split=int(metrics[paramlist[1]]),
                                max_depth=int(metrics[paramlist[2]]))

##################4.MLP超参数优化
MLP = MLPRegressor()
val=0
for i, (train_index, test_index) in enumerate(kf.split(train_x)):
    ## 用来建模的部分
    x_tra, y_tra = pd.DataFrame(train_x).iloc[train_index], pd.DataFrame(train_y).iloc[train_index]
    ## 用来做验证的部分
    x_tst, y_tst = pd.DataFrame(train_x).iloc[test_index], pd.DataFrame(train_y).iloc[test_index]
    y_tra = np.array(y_tra).ravel()  # 将二维数组转化成一维数组
    MLP.fit(x_tra, y_tra)
    predict = MLP.predict(x_tst)
    tt=r2_score(y_tst, predict)
    print(tt)
    val=val+r2_score(y_tst, predict)
print(val/n_folds)
# a.建立贝叶斯优化对象（第一个参数是优化目标函数，第二个参数是所需优化的超参数名称及其对应范围）
MLP_bo = BayesianOptimization(
    MLP_cv,
    {'hidden_layer_sizes': (1, 20),
     'max_iter': (200, 2000)})
# b.开始优化
MLP_bo.maximize(init_points=5,n_iter=20)
# c.寻找最优的几组参数做贝叶斯进阶优化
paramlist = ['hidden_layer_sizes', 'max_iter']
result = MLP_bo.res
result.sort(key=lambda x: x['target'])  # 按照target的值从小到大排列
result.reverse()
metrics = result[0]['params']
metricsave.append(metrics)
print(metrics)
MLP_model = MLPRegressor(hidden_layer_sizes=int(metrics[paramlist[0]]),
                         max_iter=int(metrics[paramlist[1]]))

##################5.ridge
ridge = Ridge()
val=0
for i, (train_index, test_index) in enumerate(kf.split(train_x)):
    ## 用来建模的部分
    x_tra, y_tra = pd.DataFrame(train_x).iloc[train_index], pd.DataFrame(train_y).iloc[train_index]
    ## 用来做验证的部分
    x_tst, y_tst = pd.DataFrame(train_x).iloc[test_index], pd.DataFrame(train_y).iloc[test_index]
    y_tra = np.array(y_tra).ravel()  # 将（135,1）的二维数组转化成一维数组
    ridge.fit(x_tra, y_tra)
    predict = ridge.predict(x_tst)
    tt=r2_score(y_tst, predict)
    print(tt)
    val=val+r2_score(y_tst, predict)
print(val/n_folds)
#a.建立贝叶斯优化对象（第一个参数是优化目标函数，第二个参数是所需优化的超参数名称及其对应范围）
ridge_bo = BayesianOptimization(
    ridge_cv,
    {'alpha': (0, 200)
    })
# b.开始优化（贝叶斯超参数优化本身也含有参数，如可以修改高斯过程的参数，其主要参数是核函数(kernel)，还有其他参数可以参考sklearn.gaussianprocess:
ridge_bo.maximize(init_points=5, n_iter=20)
# c.寻找最优的几组参数做贝叶斯进阶优化
paramlist = ['alpha']
result = ridge_bo.res
result.sort(key=lambda x: x['target'])  # 按照target的值从小到大排列
result.reverse()
metrics = result[0]['params']
metricsave.append(metrics)
print(metrics)
ridge_model = Ridge(alpha=metrics[paramlist[0]])

#####################1.5ARD
ARD = ARDRegression()
val=0
for i, (train_index, test_index) in enumerate(kf.split(train_x)):
    ## 用来建模的部分
    x_tra, y_tra = pd.DataFrame(train_x).iloc[train_index], pd.DataFrame(train_y).iloc[train_index]
    ## 用来做验证的部分
    x_tst, y_tst = pd.DataFrame(train_x).iloc[test_index], pd.DataFrame(train_y).iloc[test_index]
    y_tra = np.array(y_tra).ravel()  # 将（135,1）的二维数组转化成一维数组
    ARD.fit(x_tra, y_tra)
    predict = ARD.predict(x_tst)
    tt=r2_score(y_tst, predict)
    print(tt)
    val=val+r2_score(y_tst, predict)
print(val/n_folds)
# a.建立贝叶斯优化对象（第一个参数是优化目标函数，第二个参数是所需优化的超参数名称及其对应范围）
ard_bo = BayesianOptimization(
    ard_cv,
    {'alpha_1': (0, 1),'alpha_2': (0, 1),'lambda_1':(0,1),'lambda_2':(0,1)
    })
# b.开始优化（贝叶斯超参数优化本身也含有参数，如可以修改高斯过程的参数，其主要参数是核函数(kernel)，还有其他参数可以参考sklearn.gaussianprocess:
ard_bo.maximize(init_points=5, n_iter=20)
# c.寻找最优的几组参数做贝叶斯进阶优化
paramlist = ['alpha_1','alpha_2','lambda_1','lambda_2']
result = ard_bo.res
result.sort(key=lambda x: x['target'])  # 按照target的值从小到大排列
result.reverse()
metrics = result[0]['params']
metricsave.append(metrics)
print(metrics)
ARD_model = ARDRegression(alpha_1=metrics[paramlist[0]],alpha_2=metrics[paramlist[1]],lambda_1=metrics[paramlist[2]],lambda_2=metrics[paramlist[3]])



###超参数优化后
extendr2=[]
Q_califorenew=[]
#for clf in [rf_model, gbdt_model, MLP_model, ridge_model]:
for clf in [extra_model, gbdt_model, MLP_model, ARD_model]:
    clf.fit(train_x, train_y)
    predict = clf.predict(test_x)
    Q_califore1 = Q_forecasttest - predict  # 经校正后的预报流量
    val = r2_score(test_y, predict)
    extendr2.append(val)
    Q_califorenew.append(Q_califore1)

##超参数优化前
#Cantu
# rf_modelini = RandomForestRegressor()
# gbdt_modelini = GradientBoostingRegressor()
# MLP_modelini = MLPRegressor()
# ridge_modelini=Ridge()
#Bovisio
extra_modelini = ExtraTreesRegressor()
gbdt_modelini = GradientBoostingRegressor()
MLP_modelini = MLPRegressor()
ARD_modelini=ARDRegression()

originalr2 = []
Q_califoreold = []
#for clf in [rf_modelini,gbdt_modelini, MLP_modelini, ridge_modelini]:
for clf in [extra_modelini,gbdt_modelini, MLP_modelini, ARD_modelini]:
    clf.fit(train_x, train_y)
    predict = clf.predict(test_x)
    Q_califore2 = Q_forecasttest - predict  # 经校正后的预报流量
    val = r2_score(test_y, predict)
    originalr2.append(val)
    Q_califoreold.append(Q_califore2)

#结果输出
wb = openpyxl.Workbook()  # 计算结果导出excel
sheetname ="超参数优化后前对比"
ws = wb.create_sheet(sheetname)
Q_califorenewmean=np.zeros(len(Q_observedtest))
for k in range(len(Q_observedtest)):
    ws.cell(row=k + 1, column=1).value = Q_observedtest[k]
    ws.cell(row=k + 1, column=2).value = Q_califorenew[0][k]
    ws.cell(row=k + 1, column=3).value = Q_califorenew[1][k]
    ws.cell(row=k + 1, column=4).value = Q_califorenew[2][k]
    ws.cell(row=k + 1, column=5).value = Q_califorenew[3][k]
    Q_califorenewmean[k]=(Q_califorenew[0][k]+Q_califorenew[1][k]+Q_califorenew[2][k]+Q_califorenew[3][k])/4
    ws.cell(row=k + 1, column=6).value =Q_califorenewmean[k]
ws.cell(row=1, column=7).value=r2_score(Q_observedtest,Q_califorenew[0])
ws.cell(row=2, column=7).value=r2_score(Q_observedtest,Q_califorenew[1])
ws.cell(row=3, column=7).value=r2_score(Q_observedtest,Q_califorenew[2])
ws.cell(row=4, column=7).value=r2_score(Q_observedtest,Q_califorenew[3])
ws.cell(row=5, column=7).value=r2_score(Q_observedtest,Q_califorenewmean)

Q_califoreoldmean = np.zeros(len(Q_observedtest))
for k in range(len(Q_observedtest)):
    ws.cell(row=k + 1, column=9).value = Q_observedtest[k]
    ws.cell(row=k + 1, column=10).value = Q_califoreold[0][k]
    ws.cell(row=k + 1, column=11).value = Q_califoreold[1][k]
    ws.cell(row=k + 1, column=12).value = Q_califoreold[2][k]
    ws.cell(row=k + 1, column=13).value = Q_califoreold[3][k]
    Q_califoreoldmean[k] = (Q_califoreold[0][k] + Q_califoreold[1][k] + Q_califoreold[2][k] + Q_califoreold[3][k])/4
    ws.cell(row=k + 1, column=14).value = Q_califoreoldmean[k]
ws.cell(row=1, column=15).value=r2_score(Q_observedtest,Q_califoreold[0])
ws.cell(row=2, column=15).value=r2_score(Q_observedtest,Q_califoreold[1])
ws.cell(row=3, column=15).value=r2_score(Q_observedtest,Q_califoreold[2])
ws.cell(row=4, column=15).value=r2_score(Q_observedtest,Q_califoreold[3])
ws.cell(row=5, column=15).value=r2_score(Q_observedtest,Q_califoreoldmean)


#记录使用的超参数值
#cantu
#name=['rf','gbdt','MLP','Ridge']
#bovisio
name=['extra','gbdt','MLP','ARD']
for i in range(len(metricsave)):
    ws.cell(row=1 + i, column=17).value = name[i]
    ws.cell(row=1 + i, column=18).value=str(metricsave[i])

wb.save('filename.xlsx')



################################################# 调用Stacking函数， 得到第二层数据
train_sets = []
test_sets = []
#for clf in [rf_model,gbdt_model,MLP_model,ridge_model]:
for clf in [extra_model, gbdt_model, MLP_model, ARD_model]:
    train_set, test_set = get_stacking(clf, train_x, train_y, test_x)
    train_sets.append(train_set)
    test_sets.append(test_set)
## 把第二层数据转化为DataFrame
meta_train = pd.DataFrame(train_sets).T
meta_test = pd.DataFrame(test_sets).T
#meta_test.columns = ['rf','gbdt','MLP','Ridge']
meta_test.columns = ['extra','gbdt','MLP','ARD']
# 使用线性回归作为我们的次级分类器
dt_model =LinearRegression()
dt_model.fit(meta_train, train_y)
df_predict = pd.Series(dt_model.predict(meta_test))
df_predict1= pd.DataFrame(df_predict,columns=['stacking'])
result=meta_test.join(df_predict1)#将各回归算法的计算结果与stacking集成后的最终结果合并到result中

#将单个机器学习算法及stacking写进excel
#Cantu
#wb = openpyxl.load_workbook('filename.xlsx')#计算结果导出excel
#Bovisio
wb = openpyxl.load_workbook('filename.xlsx')#计算结果导出excel
sheetname = 'stackingcorrected'
ws = wb.create_sheet(sheetname)
#modelname=['rf','gbdt','MLP','Ridge','stacking']
modelname=['extra','gbdt','MLP','ARD','stacking']
ws.cell(row=1, column=1).value='observed'
ws.cell(row=1, column=2).value='Moloch simulated'
for i in range(len(modelname)):
    ws.cell(row=1, column=3+i).value=modelname[i]
    predict = result[modelname[i]].tolist()
    Q_califorecastest = Q_forecasttest - predict#经校正后的预报流量
    for k in range(len(Q_observedtest)):
        if i==0:
            ws.cell(row=k + 2, column=1).value = Q_observedtest[k]
            ws.cell(row=k + 2, column=2).value = Q_forecasttest[k]
        ws.cell(row=k + 2, column=3+i).value = Q_califorecastest[k]

    ws.cell(row=1, column=12 + i).value = explained_variance_score(Q_observedtest,Q_califorecastest)
    ws.cell(row=2, column=12+ i).value = mean_absolute_error(Q_observedtest, Q_califorecastest)
    ws.cell(row=3, column=12+ i).value = r2_score(Q_observedtest, Q_califorecastest)
    ws.cell(row=4, column=12 + i).value = mean_squared_error(Q_observedtest, Q_califorecastest)
ws.cell(row=1, column=3+i+1).value ='explained_variance_score';ws.cell(row=1, column=3+i+2).value =explained_variance_score(Q_observedtest,Q_forecasttest)
ws.cell(row=2, column=3+i+1).value='mean_absolute_error';ws.cell(row=2, column=3+i+2).value=mean_absolute_error(Q_observedtest,Q_forecasttest)
ws.cell(row=3, column=3+i+1).value = 'r2_score';ws.cell(row=3, column=3+i+2).value =  r2_score(Q_observedtest,Q_forecasttest)
ws.cell(row=4, column=3+i+1).value = 'mean_squared_error';ws.cell(row=4, column=3+i+2).value = mean_squared_error(Q_observedtest,Q_forecasttest)

wb.save('filename.xlsx')